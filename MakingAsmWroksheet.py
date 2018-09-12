#This script is to make ASM worksheet from invoicePackingList excel file
import openpyxl, os
import win32com.client as win32
import time

##########
#Function to delete empty rows at the end of each sheet
def cleanExcel(filename):


    wb = openpyxl.load_workbook(filename)


    for sheet in wb:
        print(sheet.title + ' max row before: ' + str(sheet.max_row))
        for i in range(12, sheet.max_row + 1):

            if sheet['I' + str(i)].value == None:
                sheet.delete_rows(i, sheet.max_row + 1 - i)
        print('max row after: ' + str(sheet.max_row))

    wb.save(filename)



#################
# Due to unknown reason, the files created must be open and save by excel before can be uploaded to ASM
# So need to run below function after file created
def ExcelSaveClose(filename):
    excel = win32.gencache.EnsureDispatch('Excel.Application') # opens Excel
    wb = excel.Workbooks.Open(os.path.join(os.getcwd(), filename))  # opens the file
    wb.Save()
    wb.Close()
    excel.Quit()

##################
# Function to choose the file and save to destFolder
def welcome():
    originFile = input('Please enter the full path of the excel file (ending with .xlsx): \n')

    destFolder = input('Please enter the folder name you want to save the gererated worksheets: \n')

    return originFile, destFolder




def WorksheetMaker():

    filename, destFolder = welcome()
#    filename = r'C:\Users\Xiaodong\Documents\GitHub\Make-ASM-template-from-Excel\invoice  packing list 217-22486085 112.xlsx'        old method, replaced by welcome function
    timestr = time.strftime("%Y%m%d-%H%M")
    destFolder = destFolder + '_' + timestr
    savePath = os.path.join(os.environ['USERPROFILE'], 'Documents', 'Worksheet', destFolder)

    if os.path.exists(savePath) == False:
        os.makedirs(savePath)
    os.chdir(savePath)


    cleanExcel(filename)

    templatename = r'C:\Users\Xiaodong\Documents\GitHub\Make-ASM-template-from-Excel\ASMtemplate.xlsx'        #ASM TEMPLATE FILE LOCATION AND name

    wb_s = openpyxl.load_workbook(filename, read_only=True, data_only=True)  #source file _s


    newFileList = []

    for sheet in wb_s:
        wb_t = openpyxl.load_workbook(templatename)               #target file _t
        sheet_t = wb_t.active

        newfilename = sheet.title
        newfilename = newfilename.replace(' ', '_')                # replace space in name
        newFileList.append(newfilename + '.xlsx')
        i = 2                                                     # from which row in template to start filling the data
        for eachrow in range (12, sheet.max_row):
            sheet_t.cell(row=i, column=5).value = 'CN'                 # column E --- CN
            sheet_t.cell(row=i, column=6).value = sheet.cell(row=eachrow, column=2).value   # col F --- B12
            sheet_t.cell(row=i, column=7).value = sheet.cell(row=eachrow, column=10).value  # col G --- J12
            sheet_t.cell(row=i, column=8).value = sheet.cell(row=eachrow, column=5).value   # col H --- E12
            sheet_t.cell(row=i, column=9).value = 100                                       # col I --- 100
            sheet_t.cell(row=i, column=10).value = 4000000                                  # col J --- 4000000
            sheet_t.cell(row=i, column=11).value = sheet.cell(row=eachrow, column=6).value   # col K --- F12
            sheet_t.cell(row=i, column=13).value = sheet.cell(row=eachrow, column=4).value    # col M --- D12
            sheet_t.cell(row=i, column=14).value = 'GBP'                                      # col N --- GBP, change to other currency if necessary
            sheet_t.cell(row=i, column=16).value = sheet.cell(row=eachrow, column=9).value    # col P --- I12
            sheet_t.cell(row=i, column=19).value = 'B'                                        # col S --- B
            sheet_t.cell(row=i, column=20).value = 0                                          # col T --- 0
            sheet_t.cell(row=i, column=24).value = 'S'                                        # col X --- S
            sheet_t.cell(row=i, column=25).value = sheet.cell(row=eachrow, column=4).value    # col Y --- D12
            sheet_t.cell(row=i, column=26).value = 'PK'                                       # col Z --- PK
            sheet_t.cell(row=i, column=27).value = r'N/M'                                     # col AA --- 'N/M'
            sheet_t.cell(row=i, column=53).value = 'Z'                                        # col BA --- 'Z'
            sheet_t.cell(row=i, column=54).value = 380                                        # col BB --- 380
            sheet_t.cell(row=i, column=55).value = sheet.cell(row=6, column=2).value          # col BC --- B6 fixed value
            sheet_t.cell(row=i, column=65).value = sheet.cell(row=6, column=2).value          # col BM --- B6 fixed value


            i += 1



        wb_t.save(newfilename + '.xlsx')

    for file in newFileList:
        ExcelSaveClose(file)


    print("%s worksheets saved successfully %s" %(str(len(wb_s.sheetnames)), savePath))
