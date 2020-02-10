from argparse import ArgumentParser
from gooey import Gooey, GooeyParser
import openpyxl, os
import win32com.client as win32
import time
import os
import json

@Gooey(program_name="Creat ASM Worksheet from InvoicePackingList")
def parse_args():
    """ Use GooeyParser to build up the arguments we will use in our script
    Save the arguments in a default json file so that we can retrieve them
    every time we run the script.
    """

    stored_args = {}
    # get the script name without the extension & use it to build up
    # the json filename
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    args_file = "{}-args.json".format(script_name)
    # Read in the prior arguments as a dictionary
    if os.path.isfile(args_file):
        with open(args_file) as data_file:
            stored_args = json.load(data_file)
    parser = GooeyParser(description='Create ASM Worksheet')
    parser.add_argument('Source_ExcelFile',
                        action='store',
                        default=stored_args.get('Source_ExcelFile'),
                        widget='FileChooser',
                        help="Source Invoice-Packing-List Excel files",
                        )
    parser.add_argument('output_folder',
                        action='store',
                        default=stored_args.get('output_folder'),
                        help="Output directory to save summary report",
                        )
    parser.add_argument('temp_file',
                        action='store',
                        default=stored_args.get('temp_file'),
                        widget='FileChooser',
                        help='Template File',
                        )
    args = parser.parse_args()
    # Store the values of the arguments so we have them next time we run
    with open(args_file, 'w') as data_file:
        # Using vars(args) returns the data as dictionary
        json.dump(vars(args), data_file)
    return args


##########
#Function to delete empty rows at the end of each sheet
def cleanExcel(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet in wb:
#        print(sheet.title + ' max row before: ' + str(sheet.max_row))
        for i in range(12, sheet.max_row + 1):
            if sheet['I' + str(i)].value == None:
                sheet.delete_rows(i, sheet.max_row + 1 - i)
#        print('max row after: ' + str(sheet.max_row))
    wb.save(filename)


###########
# Function to delete trailing spaces on HS CODE column J
def trimExcel(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet in wb:
        for row in sheet.iter_rows(min_row=12, min_col=10, max_row=sheet.max_row - 1, max_col=10):
            for cell in row:
                if type(cell.value) == str:
                    try:
                        cell.value = int(cell.value.strip())
                    except ValueError:
                        pass
    wb.save(filename)



#################
# Due to unknown reason, the files created must be open and save by excel before can be uploaded to ASM
# So need to run below function after file created
def ExcelSaveClose(filename):
    excel = win32.gencache.EnsureDispatch('Excel.Application') # opens Excel
    wb = excel.Workbooks.Open(os.path.join(os.getcwd(), filename))  # opens the file
    wb.Save()
    wb.Close()
    excel.Quit()


def WorksheetMaker(filename, destFolder, template):

#    filename, destFolder = welcome()  replaced by GUI
#    filename = r'C:\Users\Xiaodong\Documents\GitHub\Make-ASM-template-from-Excel\invoice  packing list 217-22486085 112.xlsx'        old method, replaced by welcome function
    timestr = time.strftime("%Y%m%d-%H%M")
    destFolder = destFolder + '_' + timestr
    savePath = os.path.join(os.environ['USERPROFILE'], 'Documents', 'Worksheet', destFolder)

    if os.path.exists(savePath) == False:
        os.makedirs(savePath)
    os.chdir(savePath)

    print("Cleaning the Excel File...")
    cleanExcel(filename)
    print("Deleting trailing spaces on HS CODE column...")
    trimExcel(filename)
    print("Creating Worksheet, please wait...")

#    templatename = r'C:\Users\Xiaodong\Documents\GitHub\Make-ASM-template-from-Excel\ASMtemplate.xlsx'        #ASM TEMPLATE FILE LOCATION AND name
    templatename = template


    wb_s = openpyxl.load_workbook(filename, read_only=True, data_only=True)  #source file _s


    newFileList = []

    for sheet in wb_s:
        wb_t = openpyxl.load_workbook(templatename)               #target file _t
        sheet_t = wb_t.active

        newfilename = sheet.title
        newfilename = newfilename.replace(' ', '_')                # replace space in name
        newFileList.append(newfilename + '.xlsx')
        i = 2                                                     # from which row in template to start filling the data
        for eachrow in range (12, sheet.max_row):
            sheet_t.cell(row=i, column=5).value = 'CN'  # column E --- CN
            sheet_t.cell(row=i, column=6).value = sheet.cell(row=eachrow, column=1).value  # col F --- B12, CHANGE TO A12, 10/02/2020, based on new inv/packing
            sheet_t.cell(row=i, column=7).value = sheet.cell(row=eachrow, column=8).value  # col G --- J12, change to H12, 10/02/2020
            # sheet_t.cell(row=i, column=8).value = sheet.cell(row=eachrow, column=5).value   # col H --- E12  GW COLUMN, REMOVED ON 11/12/2019 as requested by NEPTUNE
            sheet_t.cell(row=i, column=9).value = 100  # col I --- 100
            sheet_t.cell(row=i, column=10).value = 4000000  # col J --- 4000000
            sheet_t.cell(row=i, column=11).value = sheet.cell(row=eachrow, column=4).value  # col K --- F12, change to D12, 10/02/2020
            # sheet_t.cell(row=i, column=13).value = sheet.cell(row=eachrow, column=4).value    # col M --- D12
            sheet_t.cell(row=i, column=13).value = sheet.cell(row=eachrow, column=9).value  # col M --- M12, supplyment unit, change to I12, 10/02/2020
            # sheet_t.cell(row=i, column=14).value = 'GBP'                                      # col N --- GBP, change to other currency if necessary
            sheet_t.cell(row=i, column=14).value = sheet.cell(row=11, column=6).value  # col N --- H11 fixed value, change to F11, 10/02/2020
            sheet_t.cell(row=i, column=16).value = sheet.cell(row=eachrow, column=7).value  # col P --- I12, change to G12, 10/02/2020
            sheet_t.cell(row=i, column=19).value = 'B'  # col S --- B
            sheet_t.cell(row=i, column=20).value = 0  # col T --- 0
            sheet_t.cell(row=i, column=24).value = 'S'  # col X --- S
            sheet_t.cell(row=i, column=25).value = sheet.cell(row=eachrow, column=3).value  # col Y --- D12, change to C12, 10/02/2020
            sheet_t.cell(row=i, column=26).value = 'PK'  # col Z --- PK
            sheet_t.cell(row=i, column=27).value = r'N/M'  # col AA --- 'N/M'
            sheet_t.cell(row=i, column=53).value = 'Z'  # col BA --- 'Z'
            sheet_t.cell(row=i, column=54).value = 380  # col BB --- 380
            sheet_t.cell(row=i, column=55).value = sheet.cell(row=6, column=2).value  # col BC --- B6 fixed value
            sheet_t.cell(row=i, column=65).value = sheet.cell(row=6, column=2).value  # col BM --- B6 fixed value

        i += 1



        wb_t.save(newfilename + '.xlsx')
    print("Final touch...")
    for file in newFileList:
        ExcelSaveClose(file)
    print("Done!")
    print("%s worksheets saved successfully %s" %(str(len(wb_s.sheetnames)), savePath))



if __name__ == '__main__':
    conf = parse_args()
    WorksheetMaker(conf.Source_ExcelFile, conf.output_folder, conf.temp_file)
